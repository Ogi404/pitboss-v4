"""
Pitboss v4 - Brief Agent Tests

Tests for brief parsing, confidence scoring, and state machine.
"""

import pytest
from pathlib import Path
from tempfile import TemporaryDirectory
from openpyxl import Workbook

from ingest.brief_model import (
    ArticleType,
    BriefState,
    BriefKeyword,
    BriefKeywords,
    BriefSection,
    Clarification,
    BriefModel,
    BriefResult,
)
from ingest.brief_base import (
    RawBriefExtraction,
    RawKeywordGroup,
    RawSection,
    BriefParserRegistry,
    map_task_to_article_type,
)
from ingest.brief_agent import BriefAgent


# ============================================================================
# BriefModel Tests
# ============================================================================

class TestBriefKeyword:
    """Tests for BriefKeyword dataclass."""

    def test_valid_keyword_exact(self):
        """Test valid keyword with exact quantity (min=max)."""
        kw = BriefKeyword(keyword="casino bonus", min_quantity=3, max_quantity=3, group="main", confidence=0.9)
        assert kw.keyword == "casino bonus"
        assert kw.min_quantity == 3
        assert kw.max_quantity == 3
        assert kw.quantity == 3  # Backwards-compatible property
        assert kw.is_exact
        assert kw.group == "main"
        assert kw.confidence == 0.9

    def test_valid_keyword_any(self):
        """Test valid keyword with no quantity constraint."""
        kw = BriefKeyword(keyword="bonus", min_quantity=None, max_quantity=None, group="main")
        assert kw.min_quantity is None
        assert kw.max_quantity is None
        assert kw.quantity == 1  # Default via property
        assert kw.is_any

    def test_valid_keyword_range(self):
        """Test valid keyword with quantity range."""
        kw = BriefKeyword(keyword="casino", min_quantity=1, max_quantity=3, group="main")
        assert kw.min_quantity == 1
        assert kw.max_quantity == 3
        assert not kw.is_exact
        assert not kw.is_any

    def test_default_confidence(self):
        """Test default confidence is 1.0."""
        kw = BriefKeyword(keyword="test", min_quantity=1, max_quantity=1, group="main")
        assert kw.confidence == 1.0

    def test_invalid_confidence_high(self):
        """Test confidence > 1 raises error."""
        with pytest.raises(ValueError, match="Confidence must be 0-1"):
            BriefKeyword(keyword="test", min_quantity=1, max_quantity=1, group="main", confidence=1.5)

    def test_invalid_confidence_low(self):
        """Test confidence < 0 raises error."""
        with pytest.raises(ValueError, match="Confidence must be 0-1"):
            BriefKeyword(keyword="test", min_quantity=1, max_quantity=1, group="main", confidence=-0.1)

    def test_negative_min_quantity(self):
        """Test negative min_quantity raises error."""
        with pytest.raises(ValueError, match="min_quantity must be non-negative"):
            BriefKeyword(keyword="test", min_quantity=-1, max_quantity=1, group="main")

    def test_negative_max_quantity(self):
        """Test negative max_quantity raises error."""
        with pytest.raises(ValueError, match="max_quantity must be non-negative"):
            BriefKeyword(keyword="test", min_quantity=1, max_quantity=-1, group="main")

    def test_min_greater_than_max(self):
        """Test min > max raises error."""
        with pytest.raises(ValueError, match="min_quantity .* cannot exceed max_quantity"):
            BriefKeyword(keyword="test", min_quantity=5, max_quantity=2, group="main")

    def test_invalid_group(self):
        """Test invalid group raises error."""
        with pytest.raises(ValueError, match="Group must be main/support/lsi"):
            BriefKeyword(keyword="test", min_quantity=1, max_quantity=1, group="invalid")


class TestBriefKeywords:
    """Tests for BriefKeywords dataclass."""

    def test_all_keywords(self):
        """Test all_keywords property."""
        kw1 = BriefKeyword(keyword="main1", min_quantity=1, max_quantity=1, group="main")
        kw2 = BriefKeyword(keyword="support1", min_quantity=2, max_quantity=2, group="support")
        kw3 = BriefKeyword(keyword="lsi1", min_quantity=1, max_quantity=1, group="lsi")

        kws = BriefKeywords(main=(kw1,), support=(kw2,), lsi=(kw3,))
        assert len(kws.all_keywords) == 3

    def test_total_count(self):
        """Test total_count property."""
        kw1 = BriefKeyword(keyword="main1", min_quantity=3, max_quantity=3, group="main")
        kw2 = BriefKeyword(keyword="main2", min_quantity=2, max_quantity=2, group="main")

        kws = BriefKeywords(main=(kw1, kw2))
        assert kws.total_count == 5

    def test_total_count_with_any(self):
        """Test total_count with 'any' keywords counts as 1."""
        kw1 = BriefKeyword(keyword="main1", min_quantity=None, max_quantity=None, group="main")
        kw2 = BriefKeyword(keyword="main2", min_quantity=2, max_quantity=2, group="main")

        kws = BriefKeywords(main=(kw1, kw2))
        assert kws.total_count == 3  # 1 (default) + 2

    def test_min_confidence(self):
        """Test min_confidence property."""
        kw1 = BriefKeyword(keyword="main1", min_quantity=1, max_quantity=1, group="main", confidence=0.9)
        kw2 = BriefKeyword(keyword="main2", min_quantity=1, max_quantity=1, group="main", confidence=0.5)

        kws = BriefKeywords(main=(kw1, kw2))
        assert kws.min_confidence == 0.5

    def test_empty_min_confidence(self):
        """Test min_confidence with no keywords."""
        kws = BriefKeywords()
        assert kws.min_confidence == 0.0


class TestBriefSection:
    """Tests for BriefSection dataclass."""

    def test_valid_section(self):
        """Test valid section creation."""
        sec = BriefSection(heading="Introduction", word_count=200, is_required=True)
        assert sec.heading == "Introduction"
        assert sec.word_count == 200
        assert sec.is_required is True

    def test_optional_word_count(self):
        """Test section without word count."""
        sec = BriefSection(heading="Conclusion")
        assert sec.word_count is None
        assert sec.is_required is True

    def test_negative_word_count(self):
        """Test negative word count raises error."""
        with pytest.raises(ValueError, match="Word count must be non-negative"):
            BriefSection(heading="Test", word_count=-100)


class TestClarification:
    """Tests for Clarification dataclass."""

    def test_valid_clarification(self):
        """Test valid clarification creation."""
        clar = Clarification(
            field="keywords",
            question="Are these the correct keywords?",
            detected_value=["keyword1", "keyword2"],
            confidence=0.5,
        )
        assert clar.field == "keywords"
        assert clar.confidence == 0.5

    def test_with_options(self):
        """Test clarification with options."""
        clar = Clarification(
            field="article_type",
            question="What type is this?",
            detected_value="general",
            confidence=0.4,
            options=("bonus_page", "main_review", "general"),
        )
        assert len(clar.options) == 3


class TestBriefResult:
    """Tests for BriefResult dataclass."""

    def test_ready_state(self):
        """Test READY state requires brief."""
        brief = _create_mock_brief()
        result = BriefResult(state=BriefState.READY, brief=brief)
        assert result.is_ready
        assert not result.needs_input

    def test_ready_without_brief_raises(self):
        """Test READY state without brief raises error."""
        with pytest.raises(ValueError, match="READY state requires a brief"):
            BriefResult(state=BriefState.READY, brief=None)

    def test_needs_clarification_state(self):
        """Test NEEDS_CLARIFICATION state requires clarifications."""
        clar = Clarification(
            field="keywords",
            question="Confirm keywords?",
            detected_value=[],
            confidence=0.3,
        )
        result = BriefResult(
            state=BriefState.NEEDS_CLARIFICATION,
            clarifications=(clar,),
        )
        assert not result.is_ready
        assert result.needs_input

    def test_needs_clarification_without_clarifications_raises(self):
        """Test NEEDS_CLARIFICATION without clarifications raises error."""
        with pytest.raises(ValueError, match="NEEDS_CLARIFICATION state requires clarifications"):
            BriefResult(state=BriefState.NEEDS_CLARIFICATION, clarifications=())

    def test_needs_task_selection_state(self):
        """Test NEEDS_TASK_SELECTION state requires task_options."""
        result = BriefResult(
            state=BriefState.NEEDS_TASK_SELECTION,
            task_options=("Task 1", "Task 2"),
        )
        assert not result.is_ready
        assert result.needs_input

    def test_needs_task_selection_without_options_raises(self):
        """Test NEEDS_TASK_SELECTION without options raises error."""
        with pytest.raises(ValueError, match="NEEDS_TASK_SELECTION state requires task_options"):
            BriefResult(state=BriefState.NEEDS_TASK_SELECTION, task_options=())


# ============================================================================
# Task → Article Type Mapping Tests
# ============================================================================

class TestTaskToArticleTypeMapping:
    """Tests for task name to article type mapping."""

    @pytest.mark.parametrize("task_name,expected_type", [
        ("Casino Review", ArticleType.MAIN_REVIEW),
        ("full review", ArticleType.MAIN_REVIEW),
        ("site review", ArticleType.MAIN_REVIEW),
        ("Bonus Page", ArticleType.BONUS_PAGE),
        ("Welcome Offer", ArticleType.BONUS_PAGE),
        ("No Deposit Bonus", ArticleType.BONUS_PAGE),
        ("Free Spins Promotion", ArticleType.BONUS_PAGE),
        ("Mobile App Review", ArticleType.APP_REVIEW),
        ("iOS Download", ArticleType.APP_REVIEW),
        ("Android App", ArticleType.APP_REVIEW),
        ("Slot Review", ArticleType.GAME_REVIEW),
        ("Sweet Bonanza Game Review", ArticleType.GAME_REVIEW),
        ("Boxing Betting", ArticleType.SPORTS_MARKET),
        ("Football Markets", ArticleType.SPORTS_MARKET),
        ("Payment Methods", ArticleType.PAYMENTS),
        ("Banking Options", ArticleType.PAYMENTS),
        ("Registration Guide", ArticleType.REGISTRATION),
        ("Sign Up Process", ArticleType.REGISTRATION),
        ("Customer Support", ArticleType.CUSTOMER_SUPPORT),
        ("Live Chat Help", ArticleType.CUSTOMER_SUPPORT),
        ("Responsible Gaming", ArticleType.RESPONSIBLE_GAMING),
        ("Self-Exclusion", ArticleType.RESPONSIBLE_GAMING),
        ("VIP Program", ArticleType.VIP_LOYALTY),
        ("Loyalty Rewards", ArticleType.VIP_LOYALTY),
        ("Privacy Policy", ArticleType.PRIVACY_POLICY),
        ("Terms and Conditions", ArticleType.PRIVACY_POLICY),
        ("Live Casino", ArticleType.LIVE_CASINO),
        ("Live Dealer Games", ArticleType.LIVE_CASINO),
    ])
    def test_known_mappings(self, task_name, expected_type):
        """Test known task name mappings."""
        article_type, confidence = map_task_to_article_type(task_name)
        assert article_type == expected_type
        assert confidence >= 0.75

    def test_unknown_task_returns_general(self):
        """Test unknown task name returns GENERAL."""
        article_type, confidence = map_task_to_article_type("Random Content Update")
        assert article_type == ArticleType.GENERAL
        assert confidence == 0.50

    def test_empty_task_returns_general(self):
        """Test empty task name returns GENERAL with low confidence."""
        article_type, confidence = map_task_to_article_type("")
        assert article_type == ArticleType.GENERAL
        assert confidence == 0.3


# ============================================================================
# BriefAgent Tests
# ============================================================================

class TestBriefAgent:
    """Tests for BriefAgent orchestrator."""

    def test_parse_high_confidence_xlsx(self, tmp_path):
        """Test parsing a well-structured xlsx brief."""
        # Create mock xlsx
        xlsx_path = _create_mock_xlsx(tmp_path, high_confidence=True)

        agent = BriefAgent()
        result = agent.parse(xlsx_path)

        # Should be READY with high confidence
        assert result.state == BriefState.READY
        assert result.brief is not None
        assert result.brief.keywords_confidence >= 0.7

    def test_parse_low_confidence_triggers_clarification(self, tmp_path):
        """Test parsing a poorly-structured xlsx triggers clarification."""
        # Create mock xlsx with low confidence data
        xlsx_path = _create_mock_xlsx(tmp_path, high_confidence=False)

        agent = BriefAgent()
        result = agent.parse(xlsx_path)

        # Should need clarification
        assert result.state == BriefState.NEEDS_CLARIFICATION
        assert len(result.clarifications) > 0

    def test_parse_multi_task_triggers_selection(self, tmp_path):
        """Test parsing multi-task brief triggers task selection."""
        # Create mock xlsx with multiple tasks
        xlsx_path = _create_mock_xlsx(tmp_path, multi_task=True)

        agent = BriefAgent()
        result = agent.parse(xlsx_path)

        # Should need task selection
        assert result.state == BriefState.NEEDS_TASK_SELECTION
        assert len(result.task_options) > 1

    def test_confirm_clarifications(self, tmp_path):
        """Test confirming clarifications produces READY state."""
        xlsx_path = _create_mock_xlsx(tmp_path, high_confidence=False)

        agent = BriefAgent()
        result = agent.parse(xlsx_path)

        # Should need clarification
        assert result.state == BriefState.NEEDS_CLARIFICATION
        assert result.brief is not None

        # Confirm with user-provided data
        confirmations = {
            "keywords": {
                "main": [{"keyword": "casino bonus", "quantity": 3}],
                "support": [],
                "lsi": [],
            }
        }
        confirmed_result = agent.confirm_clarifications(result.brief, confirmations)

        # Keywords should now be high confidence
        assert confirmed_result.brief.keywords_confidence == 1.0

    def test_parse_with_task(self, tmp_path):
        """Test parsing multi-task brief with specific task."""
        xlsx_path = _create_mock_xlsx(tmp_path, multi_task=True)

        agent = BriefAgent()

        # First parse gets task selection
        result = agent.parse(xlsx_path)
        assert result.state == BriefState.NEEDS_TASK_SELECTION

        # Parse with selected task
        task_name = result.task_options[0]
        result = agent.parse_with_task(xlsx_path, task_name)

        # Should proceed (may still need clarification for other fields)
        assert result.state in (BriefState.READY, BriefState.NEEDS_CLARIFICATION)


class TestBriefParserRegistry:
    """Tests for parser registry."""

    def test_xlsx_parser_registered(self):
        """Test xlsx parser is auto-registered."""
        formats = BriefParserRegistry.list_formats()
        assert "xlsx" in formats

    def test_detect_xlsx(self, tmp_path):
        """Test auto-detection of xlsx files."""
        xlsx_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(xlsx_path)
        wb.close()

        parser = BriefParserRegistry.detect_and_get(xlsx_path)
        assert parser.get_format_name() == "xlsx"

    def test_unknown_format_raises(self, tmp_path):
        """Test unknown format raises error."""
        txt_path = tmp_path / "test.txt"
        txt_path.write_text("test")

        with pytest.raises(ValueError, match="No parser found"):
            BriefParserRegistry.detect_and_get(txt_path)


# ============================================================================
# Xlsx Parser Tests
# ============================================================================

class TestXlsxParser:
    """Tests for Excel brief parser."""

    def test_extract_keyword_table(self, tmp_path):
        """Test extracting keywords from table format."""
        xlsx_path = _create_keyword_table_xlsx(tmp_path)

        parser = BriefParserRegistry.get("xlsx")
        raw = parser.extract(xlsx_path)

        assert raw.has_keywords
        assert raw.keywords_confidence >= 0.9

        # Check keywords extracted (format is now (keyword, min_qty, max_qty))
        all_kws = []
        for group in raw.keyword_groups:
            all_kws.extend([kw_tuple[0] for kw_tuple in group.keywords])

        assert "casino bonus" in all_kws
        assert "free spins" in all_kws

    def test_extract_section_table(self, tmp_path):
        """Test extracting sections from table format."""
        xlsx_path = _create_section_table_xlsx(tmp_path)

        parser = BriefParserRegistry.get("xlsx")
        raw = parser.extract(xlsx_path)

        assert len(raw.sections) > 0
        assert raw.sections_confidence >= 0.9

        headings = [s.heading for s in raw.sections]
        assert "Introduction" in headings

    def test_extract_meta_fields(self, tmp_path):
        """Test extracting meta fields."""
        xlsx_path = _create_meta_xlsx(tmp_path)

        parser = BriefParserRegistry.get("xlsx")
        raw = parser.extract(xlsx_path)

        assert raw.brand_name == "TestBrand"
        assert raw.target_word_count == 2000
        assert raw.task_name == "Bonus Page"


# ============================================================================
# Helper Functions
# ============================================================================

def _create_mock_brief() -> BriefModel:
    """Create a mock BriefModel for testing."""
    kw = BriefKeyword(keyword="test", min_quantity=1, max_quantity=1, group="main", confidence=0.9)
    keywords = BriefKeywords(main=(kw,))
    section = BriefSection(heading="Introduction", word_count=200)

    return BriefModel(
        keywords=keywords,
        keywords_confidence=0.9,
        sections=(section,),
        sections_confidence=0.9,
        target_word_count=2000,
        word_count_confidence=0.9,
        task_name="Bonus Page",
        article_type=ArticleType.BONUS_PAGE,
        article_type_confidence=0.9,
        locale="en-CA",
        market="CA",
        locale_confidence=0.9,
        brand_name="TestBrand",
        source_path="/path/to/brief.xlsx",
        source_format="xlsx",
    )


def _create_mock_xlsx(tmp_path: Path, high_confidence: bool = True, multi_task: bool = False) -> Path:
    """Create a mock xlsx brief file."""
    xlsx_path = tmp_path / "brief.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Brief"

    if high_confidence:
        # Well-structured brief
        ws["A1"] = "Main Keywords"
        ws["B1"] = "Quantity"
        ws["A2"] = "casino bonus"
        ws["B2"] = 3
        ws["A3"] = "free spins"
        ws["B3"] = 2

        ws["D1"] = "Section"
        ws["E1"] = "Word Count"
        ws["D2"] = "Introduction"
        ws["E2"] = 200
        ws["D3"] = "Bonus Overview"
        ws["E3"] = 500

        ws["G1"] = "Brand"
        ws["H1"] = "TestBrand"
        ws["G2"] = "Word Count"
        ws["H2"] = 2000
        ws["G3"] = "Task"
        ws["H3"] = "Bonus Page"
    else:
        # Poorly structured - keywords in unclear format
        ws["A1"] = "Notes"
        ws["A2"] = "Some keywords maybe: bonus, casino"

    if multi_task:
        # Add multiple tasks
        ws["G3"] = "Task"
        ws["H3"] = "Task 1"

        ws2 = wb.create_sheet("Task 2")
        ws2["A1"] = "Task"
        ws2["B1"] = "Task 2"

    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def _create_keyword_table_xlsx(tmp_path: Path) -> Path:
    """Create xlsx with well-structured keyword table."""
    xlsx_path = tmp_path / "keywords.xlsx"
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Main Keywords"
    ws["B1"] = "Quantity"
    ws["A2"] = "casino bonus"
    ws["B2"] = 3
    ws["A3"] = "free spins"
    ws["B3"] = 2
    ws["A4"] = "welcome offer"
    ws["B4"] = 2

    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def _create_section_table_xlsx(tmp_path: Path) -> Path:
    """Create xlsx with section structure table."""
    xlsx_path = tmp_path / "sections.xlsx"
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Section"
    ws["B1"] = "Word Count"
    ws["A2"] = "Introduction"
    ws["B2"] = 200
    ws["A3"] = "Bonus Overview"
    ws["B3"] = 500
    ws["A4"] = "Terms and Conditions"
    ws["B4"] = 300
    ws["A5"] = "Conclusion"
    ws["B5"] = 150

    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def _create_meta_xlsx(tmp_path: Path) -> Path:
    """Create xlsx with meta fields."""
    xlsx_path = tmp_path / "meta.xlsx"
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Brand"
    ws["B1"] = "TestBrand"
    ws["A2"] = "Word Count"
    ws["B2"] = 2000
    ws["A3"] = "Task"
    ws["B3"] = "Bonus Page"
    ws["A4"] = "Market"
    ws["B4"] = "CA"
    ws["A5"] = "Locale"
    ws["B5"] = "en-CA"

    wb.save(xlsx_path)
    wb.close()
    return xlsx_path
