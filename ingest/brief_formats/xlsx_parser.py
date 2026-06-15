"""
Pitboss v4 - Excel Brief Parser

Parses .xlsx brief files into RawBriefExtraction with confidence scoring.

Handles format variants:
- Single-tab: All data on one sheet
- Multi-tab: Separate tabs for keywords, sections, meta
- Multi-task: Multiple tasks/articles in same workbook
"""

from __future__ import annotations
from pathlib import Path
from typing import Union, Optional, Any
import re

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from ingest.brief_base import (
    BriefParser,
    register_brief_parser,
    RawBriefExtraction,
    RawKeywordGroup,
    RawSection,
    RawLink,
    is_metadata_label,
    _parse_quantity_range,
    clean_keyword,
)


# Keyword table header patterns (high confidence 0.9+)
KEYWORD_TABLE_HEADERS = [
    (re.compile(r"main\s*keywords?", re.IGNORECASE), re.compile(r"(qty|quantity|count|#)", re.IGNORECASE)),
    (re.compile(r"primary\s*keywords?", re.IGNORECASE), re.compile(r"(qty|quantity|count|#)", re.IGNORECASE)),
    (re.compile(r"main\s*kw", re.IGNORECASE), re.compile(r"(qty|quantity|count)", re.IGNORECASE)),
    (re.compile(r"keywords?", re.IGNORECASE), re.compile(r"(qty|quantity|count|#)", re.IGNORECASE)),
]

# Keyword column hints for medium confidence (0.7-0.9)
KEYWORD_COLUMN_HINTS = [
    re.compile(r"keyword", re.IGNORECASE),
    re.compile(r"\bkw\b", re.IGNORECASE),
    re.compile(r"\bmain\b", re.IGNORECASE),
    re.compile(r"support", re.IGNORECASE),
    re.compile(r"\blsi\b", re.IGNORECASE),
    re.compile(r"secondary", re.IGNORECASE),
]

# Section/structure table headers
SECTION_TABLE_HEADERS = [
    (re.compile(r"section|heading|structure", re.IGNORECASE), re.compile(r"word\s*count|words?|length", re.IGNORECASE)),
    (re.compile(r"outline|content", re.IGNORECASE), re.compile(r"word\s*count|words?|length", re.IGNORECASE)),
]

# Meta field patterns
META_PATTERNS = {
    "brand": re.compile(r"brand|client|website|site\s*name", re.IGNORECASE),
    "locale": re.compile(r"locale|language|region", re.IGNORECASE),
    "market": re.compile(r"market|country|geo|target", re.IGNORECASE),
    "word_count": re.compile(r"(total\s*)?(word\s*count|words?|length|target\s*length)", re.IGNORECASE),
    "task": re.compile(r"task|topic|article\s*type|content\s*type|type", re.IGNORECASE),
}

# Group name patterns for categorizing keywords
GROUP_PATTERNS = {
    "main": re.compile(r"\bmain\b|primary|focus", re.IGNORECASE),
    "support": re.compile(r"support|secondary|additional", re.IGNORECASE),
    "lsi": re.compile(r"\blsi\b|latent|semantic|related", re.IGNORECASE),
}

# Link extraction patterns
LINK_ANCHOR_PATTERN = re.compile(r"link\s*\d*\s*[—\-–]?\s*anchor", re.IGNORECASE)
LINK_URL_PATTERN = re.compile(r"link\s*\d*\s*[—\-–]?\s*url", re.IGNORECASE)
LINK_HEADER_PATTERN = re.compile(r"^link\s*\d*$", re.IGNORECASE)


def _cell_value(cell: Optional[Cell]) -> Optional[str]:
    """Get string value from cell, handling None and empty."""
    if cell is None or cell.value is None:
        return None
    value = str(cell.value).strip()
    return value if value else None


def _parse_quantity(value: Any) -> int:
    """Parse quantity from cell value."""
    if value is None:
        return 1  # Default quantity
    if isinstance(value, (int, float)):
        return max(1, int(value))
    # Try to extract number from string
    text = str(value).strip()
    match = re.search(r"(\d+)", text)
    if match:
        return max(1, int(match.group(1)))
    return 1


def _infer_group_name(header: str) -> str:
    """Infer keyword group name from header text."""
    for group_name, pattern in GROUP_PATTERNS.items():
        if pattern.search(header):
            return group_name
    return "main"  # Default to main


@register_brief_parser
class XlsxBriefParser(BriefParser):
    """Parser for Excel (.xlsx) brief files."""

    def get_format_name(self) -> str:
        return "xlsx"

    def can_parse(self, source: Union[Path, str]) -> bool:
        if isinstance(source, str):
            source = Path(source)
        return source.suffix.lower() in (".xlsx", ".xls")

    def extract(self, source: Union[Path, str]) -> RawBriefExtraction:
        """Extract raw data from Excel brief."""
        if isinstance(source, str):
            source = Path(source)

        wb = load_workbook(str(source), data_only=True)

        extraction = RawBriefExtraction(
            source_path=str(source),
            source_format="xlsx",
            raw_data={"sheets": [s for s in wb.sheetnames]},
        )

        # Detect multi-task briefs
        tasks = self._detect_tasks(wb)
        extraction.tasks = tasks

        # Find and extract keywords
        keyword_groups, kw_confidence = self._extract_keywords(wb)
        extraction.keyword_groups = keyword_groups
        extraction.keywords_confidence = kw_confidence

        # Find and extract sections
        sections, sec_confidence = self._extract_sections(wb)
        extraction.sections = sections
        extraction.sections_confidence = sec_confidence

        # Extract meta fields
        meta = self._extract_meta(wb)
        extraction.brand_name = meta.get("brand")
        extraction.brand_confidence = meta.get("brand_confidence", 0.5)
        extraction.locale = meta.get("locale")
        extraction.market = meta.get("market")
        extraction.locale_confidence = meta.get("locale_confidence", 0.5)
        extraction.target_word_count = meta.get("word_count")
        extraction.word_count_confidence = meta.get("word_count_confidence", 0.5)
        extraction.task_name = meta.get("task")
        extraction.task_name_confidence = meta.get("task_confidence", 0.5)

        # Extract links
        links, links_confidence = self._extract_links(wb)
        extraction.links = links
        extraction.links_confidence = links_confidence

        wb.close()
        return extraction

    def _detect_tasks(self, wb) -> list[str]:
        """Detect task names for multi-task briefs, with row-above fallback."""
        tasks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Look for task/topic column or row
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
                for cell in row:
                    val = _cell_value(cell)
                    if val and META_PATTERNS["task"].search(val):
                        # Check adjacent cells for task value
                        next_col = cell.column + 1
                        task_val = None
                        if next_col <= ws.max_column:
                            task_cell = ws.cell(row=cell.row, column=next_col)
                            task_val = _cell_value(task_cell)

                        # Row-above fallback if adjacent cell is empty
                        if not task_val and cell.row > 1:
                            above_cell = ws.cell(row=cell.row - 1, column=next_col)
                            task_val = _cell_value(above_cell)

                        if task_val and task_val not in tasks:
                            tasks.append(task_val)

        # If no explicit tasks found, use sheet names as tasks (if meaningful)
        if not tasks:
            meaningful_sheets = [
                s for s in wb.sheetnames
                if s.lower() not in ("sheet1", "data", "info", "meta", "keywords")
            ]
            if len(meaningful_sheets) > 1:
                tasks = meaningful_sheets

        return tasks

    def _extract_keywords(self, wb) -> tuple[list[RawKeywordGroup], float]:
        """Extract keywords from workbook."""
        keyword_groups = []
        best_confidence = 0.0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            groups, confidence = self._find_keywords_in_sheet(ws)
            if groups:
                keyword_groups.extend(groups)
                best_confidence = max(best_confidence, confidence)

        # If no structured keywords found, try inline parsing
        if not keyword_groups:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                groups, confidence = self._parse_inline_keywords(ws)
                if groups:
                    keyword_groups.extend(groups)
                    best_confidence = max(best_confidence, confidence)

        # Deduplicate keywords within each group
        keyword_groups = self._deduplicate_keyword_groups(keyword_groups)

        return keyword_groups, best_confidence

    def _deduplicate_keyword_groups(self, groups: list[RawKeywordGroup]) -> list[RawKeywordGroup]:
        """Deduplicate keywords across groups, keeping highest confidence version."""
        # Track seen keywords by (keyword, group_name)
        seen = {}  # (keyword_lower, group_name) -> (keyword, min_qty, max_qty, confidence)

        for group in groups:
            for kw_tuple in group.keywords:
                keyword = kw_tuple[0]
                min_qty = kw_tuple[1] if len(kw_tuple) > 1 else None
                max_qty = kw_tuple[2] if len(kw_tuple) > 2 else None

                key = (keyword.lower(), group.group_name)

                if key not in seen:
                    seen[key] = (keyword, min_qty, max_qty, group.confidence)
                else:
                    # Keep version with higher confidence, or with defined quantities
                    existing = seen[key]
                    if group.confidence > existing[3]:
                        seen[key] = (keyword, min_qty, max_qty, group.confidence)
                    elif group.confidence == existing[3]:
                        # If same confidence, prefer version with defined quantities
                        if min_qty is not None and existing[1] is None:
                            seen[key] = (keyword, min_qty, max_qty, group.confidence)

        # Rebuild groups
        group_keywords = {}  # group_name -> list of (keyword, min_qty, max_qty)
        group_confidence = {}  # group_name -> max_confidence

        for (kw_lower, group_name), (keyword, min_qty, max_qty, conf) in seen.items():
            if group_name not in group_keywords:
                group_keywords[group_name] = []
                group_confidence[group_name] = 0.0
            group_keywords[group_name].append((keyword, min_qty, max_qty))
            group_confidence[group_name] = max(group_confidence[group_name], conf)

        # Create new groups
        result = []
        for group_name, keywords in group_keywords.items():
            result.append(RawKeywordGroup(
                keywords=keywords,
                group_name=group_name,
                confidence=group_confidence[group_name],
            ))

        return result

    def _find_keywords_in_sheet(self, ws: Worksheet) -> tuple[list[RawKeywordGroup], float]:
        """Find structured keyword tables in a worksheet."""
        groups = []
        confidence = 0.0

        # Look for keyword table headers (scan up to row 100 for briefs with long section lists)
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, max_col=15), start=1):
            for col_idx, cell in enumerate(row, start=1):
                val = _cell_value(cell)
                if not val:
                    continue

                # Check for high-confidence table header pattern
                for kw_pattern, qty_pattern in KEYWORD_TABLE_HEADERS:
                    if kw_pattern.search(val):
                        # Look for quantity column
                        qty_col = None
                        for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                            next_val = _cell_value(ws.cell(row=row_idx, column=next_col))
                            if next_val and qty_pattern.search(next_val):
                                qty_col = next_col
                                break

                        # Extract keywords from rows below
                        keywords = []
                        for data_row in range(row_idx + 1, min(row_idx + 100, ws.max_row + 1)):
                            kw_val = _cell_value(ws.cell(row=data_row, column=col_idx))
                            if not kw_val:
                                # Stop on empty row
                                if not any(_cell_value(ws.cell(row=data_row, column=c))
                                          for c in range(col_idx, min(col_idx + 5, ws.max_column + 1))):
                                    break
                                continue

                            # Filter out metadata labels
                            if is_metadata_label(kw_val):
                                continue

                            # Clean keyword (strip parenthetical translations)
                            kw_clean = clean_keyword(kw_val)
                            if not kw_clean or is_metadata_label(kw_clean):
                                continue

                            # Parse quantity as min/max
                            min_qty, max_qty = None, None
                            if qty_col:
                                qty_value = ws.cell(row=data_row, column=qty_col).value
                                min_qty, max_qty = _parse_quantity_range(qty_value)

                            keywords.append((kw_clean, min_qty, max_qty))

                        if keywords:
                            group_name = _infer_group_name(val)
                            groups.append(RawKeywordGroup(
                                keywords=keywords,
                                group_name=group_name,
                                confidence=0.95 if qty_col else 0.80,
                            ))
                            confidence = max(confidence, 0.95 if qty_col else 0.80)

        # Also check for column-hint based detection
        if not groups:
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                header = _cell_value(ws.cell(row=1, column=col_idx))
                if not header:
                    continue

                for hint in KEYWORD_COLUMN_HINTS:
                    if hint.search(header):
                        keywords = []
                        for row_idx in range(2, min(ws.max_row + 1, 100)):
                            val = _cell_value(ws.cell(row=row_idx, column=col_idx))
                            if val and not is_metadata_label(val):
                                kw_clean = clean_keyword(val)
                                if kw_clean and not is_metadata_label(kw_clean):
                                    keywords.append((kw_clean, None, None))  # No quantity constraint

                        if keywords:
                            group_name = _infer_group_name(header)
                            groups.append(RawKeywordGroup(
                                keywords=keywords,
                                group_name=group_name,
                                confidence=0.75,
                            ))
                            confidence = max(confidence, 0.75)
                        break

        return groups, confidence

    def _parse_inline_keywords(self, ws: Worksheet) -> tuple[list[RawKeywordGroup], float]:
        """Parse inline keyword patterns (comma-separated, etc.)."""
        groups = []
        confidence = 0.0

        # Look for cells with keyword-like content
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=10):
            for cell in row:
                val = _cell_value(cell)
                if not val:
                    continue

                # Check for "Keywords: word1, word2, word3" pattern
                match = re.match(r"(main\s*)?keywords?\s*[:=]\s*(.+)", val, re.IGNORECASE)
                if match:
                    keyword_text = match.group(2)
                    keywords = self._parse_keyword_list(keyword_text)
                    if keywords:
                        # Infer group from prefix if present
                        group_name = "main"
                        prefix = match.group(1)
                        if prefix:
                            group_name = _infer_group_name(prefix)

                        groups.append(RawKeywordGroup(
                            keywords=keywords,
                            group_name=group_name,
                            confidence=0.65,
                        ))
                        confidence = max(confidence, 0.65)

        return groups, confidence

    def _parse_keyword_list(self, text: str) -> list[tuple[str, Optional[int], Optional[int]]]:
        """Parse a comma/semicolon/slash-separated keyword list with optional quantities."""
        keywords = []

        # Split on comma, semicolon, OR forward slash
        parts = re.split(r"[,;/]", text)

        for part in parts:
            part = part.strip()
            if not part:
                continue

            # Filter out metadata labels
            if is_metadata_label(part):
                continue

            # Check for quantity suffix: "keyword (3x)" or "keyword x3" or "keyword - 3"
            qty_match = re.match(r"(.+?)\s*[\(\-]\s*(\d+)\s*[x×]?\s*\)?$", part)
            if qty_match:
                kw = clean_keyword(qty_match.group(1).strip())
                qty = int(qty_match.group(2))
                if kw and not is_metadata_label(kw):
                    keywords.append((kw, qty, qty))  # Exact quantity
            else:
                qty_match = re.match(r"(.+?)\s*[x×]\s*(\d+)$", part)
                if qty_match:
                    kw = clean_keyword(qty_match.group(1).strip())
                    qty = int(qty_match.group(2))
                    if kw and not is_metadata_label(kw):
                        keywords.append((kw, qty, qty))  # Exact quantity
                else:
                    kw = clean_keyword(part)
                    if kw and not is_metadata_label(kw):
                        keywords.append((kw, None, None))  # No quantity constraint

        return keywords

    def _extract_sections(self, wb) -> tuple[list[RawSection], float]:
        """Extract section/structure from workbook."""
        sections = []
        best_confidence = 0.0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_sections, confidence = self._find_sections_in_sheet(ws)
            if sheet_sections:
                sections.extend(sheet_sections)
                best_confidence = max(best_confidence, confidence)

        return sections, best_confidence

    def _find_sections_in_sheet(self, ws: Worksheet) -> tuple[list[RawSection], float]:
        """Find section/structure table in a worksheet."""
        sections = []
        confidence = 0.0

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, max_col=10), start=1):
            for col_idx, cell in enumerate(row, start=1):
                val = _cell_value(cell)
                if not val:
                    continue

                # Check for section table header pattern
                for sec_pattern, wc_pattern in SECTION_TABLE_HEADERS:
                    if sec_pattern.search(val):
                        # Look for word count column
                        wc_col = None
                        for next_col in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                            next_val = _cell_value(ws.cell(row=row_idx, column=next_col))
                            if next_val and wc_pattern.search(next_val):
                                wc_col = next_col
                                break

                        # Extract sections from rows below
                        for data_row in range(row_idx + 1, min(row_idx + 50, ws.max_row + 1)):
                            sec_val = _cell_value(ws.cell(row=data_row, column=col_idx))
                            if not sec_val:
                                # Stop on empty row
                                if not any(_cell_value(ws.cell(row=data_row, column=c))
                                          for c in range(col_idx, min(col_idx + 5, ws.max_column + 1))):
                                    break
                                continue

                            word_count = None
                            if wc_col:
                                wc_val = ws.cell(row=data_row, column=wc_col).value
                                if wc_val:
                                    word_count = _parse_quantity(wc_val)

                            sections.append(RawSection(
                                heading=sec_val,
                                word_count=word_count,
                                confidence=0.95 if wc_col else 0.75,
                            ))
                            confidence = max(confidence, 0.95 if wc_col else 0.75)

                        if sections:
                            return sections, confidence

        return sections, confidence

    def _extract_meta(self, wb) -> dict[str, Any]:
        """Extract meta fields (brand, locale, word count, etc.)."""
        meta = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Scan for key-value pairs
            for row in ws.iter_rows(min_row=1, max_row=30, max_col=10):
                for cell in row:
                    val = _cell_value(cell)
                    if not val:
                        continue

                    # Check against meta patterns
                    for field_name, pattern in META_PATTERNS.items():
                        if pattern.search(val):
                            # Get value from adjacent cell
                            next_col = cell.column + 1
                            if next_col <= ws.max_column:
                                value_cell = ws.cell(row=cell.row, column=next_col)
                                field_val = _cell_value(value_cell)
                                if field_val and field_name not in meta:
                                    if field_name == "word_count":
                                        meta[field_name] = _parse_quantity(field_val)
                                        meta[f"{field_name}_confidence"] = 0.90
                                    else:
                                        meta[field_name] = field_val
                                        meta[f"{field_name}_confidence"] = 0.85

                    # Also check for inline patterns: "Brand: Vave"
                    for field_name, pattern in META_PATTERNS.items():
                        match = re.match(rf"{pattern.pattern}\s*[:=]\s*(.+)", val, re.IGNORECASE)
                        if match and field_name not in meta:
                            field_val = match.group(1).strip()
                            if field_name == "word_count":
                                meta[field_name] = _parse_quantity(field_val)
                                meta[f"{field_name}_confidence"] = 0.80
                            else:
                                meta[field_name] = field_val
                                meta[f"{field_name}_confidence"] = 0.75

        return meta

    def _extract_links(self, wb) -> tuple[list[RawLink], float]:
        """Extract link specifications (anchor/URL pairs) from workbook."""
        links = []
        confidence = 0.0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Find anchor/url patterns
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, max_col=10), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    val = _cell_value(cell)
                    if not val:
                        continue

                    # Check for anchor pattern: "Link 1 - Anchor" or "Link 1 — Anchor"
                    if LINK_ANCHOR_PATTERN.match(val):
                        # Get anchor text from adjacent cell
                        anchor_val = None
                        if col_idx < ws.max_column:
                            anchor_val = _cell_value(ws.cell(row=row_idx, column=col_idx + 1))

                        if anchor_val:
                            # Look for corresponding URL (usually row below or nearby)
                            url_val = None

                            # Check row below at same column
                            if row_idx < ws.max_row:
                                below_label = _cell_value(ws.cell(row=row_idx + 1, column=col_idx))
                                if below_label and LINK_URL_PATTERN.match(below_label):
                                    url_val = _cell_value(ws.cell(row=row_idx + 1, column=col_idx + 1))

                            # Determine link type
                            link_type = "external"
                            if url_val:
                                # Internal if relative path or no protocol
                                if url_val.startswith("/") or (not url_val.startswith("http") and "/" in url_val):
                                    link_type = "internal"

                            links.append(RawLink(
                                anchor=anchor_val,
                                url=url_val or "",
                                link_type=link_type,
                                confidence=0.85 if url_val else 0.60,
                            ))
                            confidence = max(confidence, 0.85 if url_val else 0.60)

        return links, confidence
